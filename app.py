from flask import Flask, render_template, request, redirect, url_for, send_file, session
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import os
import csv
from io import StringIO, BytesIO
from datetime import datetime
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import matplotlib.pyplot as plt
import numpy as np

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

app = Flask(__name__)
basedir = os.path.abspath(os.path.dirname(__file__))
db_url = os.environ.get('DATABASE_URL') or 'sqlite:///' + os.path.join(basedir, 'feedback.db')
app.config['SQLALCHEMY_DATABASE_URI'] = db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')
db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Please log in to access this page.'


class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(20), nullable=False, default='student')  # 'student' or 'admin'
    
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)
    
    def check_password(self, password):
        return check_password_hash(self.password_hash, password)


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


class Faculty(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), unique=True, nullable=False)
    department = db.Column(db.String(120))
    
    def __repr__(self):
        return f'<Faculty {self.name}>'


class Feedback(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_name = db.Column(db.String(120))
    faculty_id = db.Column(db.Integer, db.ForeignKey('faculty.id'), nullable=False)
    faculty = db.relationship('Faculty', backref='feedbacks')
    
    # Questions with 10-point scale
    q1_teaching_quality = db.Column(db.Integer)  # Teaching quality
    q2_course_content = db.Column(db.Integer)    # Course content clarity
    q3_communication = db.Column(db.Integer)     # Communication skills
    q4_feedback_quality = db.Column(db.Integer)  # Quality of feedback
    q5_subject_knowledge = db.Column(db.Integer) # Subject knowledge
    
    overall_rating = db.Column(db.Integer)  # Overall rating (1-10)
    comments = db.Column(db.Text)
    created_at = db.Column(db.DateTime, server_default=db.func.current_timestamp())


def create_tables():
    db.create_all()

# Ensure tables exist when the app is started. Some Flask runtimes or versions
# may not expose the `before_first_request` decorator on the app object,
# so explicitly create tables using an application context which works
# regardless of how the app is executed (flask run, gunicorn, direct run).
try:
    with app.app_context():
        create_tables()
except RuntimeError:
    # If the app context cannot be pushed at import time, skip; tables will
    # be created on the first request or when the app context becomes available.
    pass


@app.route('/', methods=['GET', 'POST'])
def index():
    faculties = Faculty.query.all()
    
    if request.method == 'POST':
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        
        faculty_id = request.form.get('faculty_id')
        student_name = request.form.get('student_name')
        
        # Validate inputs
        try:
            q1 = int(request.form.get('q1_teaching_quality') or 0)
            q2 = int(request.form.get('q2_course_content') or 0)
            q3 = int(request.form.get('q3_communication') or 0)
            q4 = int(request.form.get('q4_feedback_quality') or 0)
            q5 = int(request.form.get('q5_subject_knowledge') or 0)
            overall = int(request.form.get('overall_rating') or 0)
        except ValueError:
            return render_template('index.html', faculties=faculties, error='Invalid rating value')
        
        # Validate all ratings are between 1-10
        ratings = [q1, q2, q3, q4, q5, overall]
        if not faculty_id or not all(1 <= r <= 10 for r in ratings):
            return render_template('index.html', faculties=faculties, error='Please provide all ratings (1-10)')
        
        comments = request.form.get('comments', '')
        
        fb = Feedback(
            student_name=student_name or 'Anonymous',
            faculty_id=int(faculty_id),
            q1_teaching_quality=q1,
            q2_course_content=q2,
            q3_communication=q3,
            q4_feedback_quality=q4,
            q5_subject_knowledge=q5,
            overall_rating=overall,
            comments=comments
        )
        db.session.add(fb)
        db.session.commit()
        return redirect(url_for('thankyou'))
    
    if not current_user.is_authenticated:
        return redirect(url_for('login'))
    
    return render_template('index.html', faculties=faculties)


@app.route('/thankyou')
def thankyou():
    return render_template('thankyou.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password):
            login_user(user)
            next_page = request.args.get('next')
            if next_page:
                return redirect(next_page)
            return redirect(url_for('index') if user.role == 'student' else url_for('admin'))
        else:
            return render_template('login.html', error='Invalid username or password')
    
    return render_template('login.html')


@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if User.query.filter_by(username=username).first():
            return render_template('register.html', error='Username already exists')
        
        user = User(username=username, role='student')
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        return redirect(url_for('login'))
    
    return render_template('register.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))


@app.route('/admin')
def admin():
    if not current_user.is_authenticated or current_user.role != 'admin':
        return redirect(url_for('login'))
    
    feedbacks = Feedback.query.order_by(Feedback.created_at.desc()).all()
    from sqlalchemy import func
    stats = db.session.query(
        Faculty.name,
        func.count(Feedback.id).label('count'),
        func.avg(Feedback.overall_rating).label('avg_rating'),
    ).join(Feedback).group_by(Faculty.id, Faculty.name).all()
    return render_template('admin.html', feedbacks=feedbacks, stats=stats)


@app.route('/analytics')
@login_required
def analytics():
    if current_user.role != 'admin':
        return redirect(url_for('index'))
    
    feedbacks = Feedback.query.all()
    if not feedbacks:
        return render_template('analytics.html', error='No feedback data available')
    
    # Prepare data for analysis
    ratings = np.array([f.overall_rating for f in feedbacks if f.overall_rating])
    faculties = [f.faculty.name for f in feedbacks]
    unique_faculties = list(set(faculties))
    
    return render_template('analytics.html', feedbacks=feedbacks, unique_faculties=unique_faculties)


@app.route('/chart/rating-distribution')
@login_required
def chart_rating_distribution():
    if current_user.role != 'admin':
        return redirect(url_for('login'))
    
    feedbacks = Feedback.query.all()
    ratings = [f.overall_rating for f in feedbacks if f.overall_rating]
    
    if not ratings:
        return "No data", 404
    
    # Create rating distribution chart (1-10 scale)
    fig, ax = plt.subplots(figsize=(12, 6))
    rating_counts = np.bincount(ratings, minlength=1)
    ax.bar(range(1, len(rating_counts) + 1), rating_counts, color='#238636', edgecolor='black')
    ax.set_xlabel('Overall Rating', fontsize=12)
    ax.set_ylabel('Count', fontsize=12)
    ax.set_title('Overall Rating Distribution (1-10)', fontsize=14, fontweight='bold')
    ax.set_xticks(range(1, 11))
    ax.grid(axis='y', alpha=0.3)
    
    # Save to BytesIO
    img = BytesIO()
    fig.savefig(img, format='png', bbox_inches='tight')
    img.seek(0)
    plt.close(fig)
    
    return send_file(img, mimetype='image/png')


@app.route('/chart/faculty-ratings')
@login_required
def chart_faculty_ratings():
    if current_user.role != 'admin':
        return redirect(url_for('login'))
    
    feedbacks = Feedback.query.all()
    if not feedbacks:
        return "No data", 404
    
    # Group ratings by faculty
    faculty_ratings = {}
    for f in feedbacks:
        fac_name = f.faculty.name
        if fac_name not in faculty_ratings:
            faculty_ratings[fac_name] = []
        if f.overall_rating:
            faculty_ratings[fac_name].append(f.overall_rating)
    
    # Calculate average for each faculty
    faculties = list(faculty_ratings.keys())
    avg_ratings = [np.mean(faculty_ratings[fac]) for fac in faculties]
    
    fig, ax = plt.subplots(figsize=(12, 6))
    colors = plt.cm.RdYlGn(np.linspace(0.3, 0.9, len(faculties)))
    bars = ax.bar(faculties, avg_ratings, color=colors, edgecolor='black')
    ax.set_ylabel('Average Rating', fontsize=12)
    ax.set_title('Average Overall Rating by Faculty (out of 10)', fontsize=14, fontweight='bold')
    ax.set_ylim(0, 10.5)
    ax.grid(axis='y', alpha=0.3)
    plt.xticks(rotation=45, ha='right')
    
    # Add value labels on bars
    for bar in bars:
        height = bar.get_height()
        ax.text(bar.get_x() + bar.get_width()/2., height,
                f'{height:.2f}', ha='center', va='bottom')
    
    img = BytesIO()
    fig.savefig(img, format='png', bbox_inches='tight')
    img.seek(0)
    plt.close(fig)
    
    return send_file(img, mimetype='image/png')


@app.route('/chart/feedback-count')
@login_required
def chart_feedback_count():
    if current_user.role != 'admin':
        return redirect(url_for('login'))
    
    feedbacks = Feedback.query.all()
    if not feedbacks:
        return "No data", 404
    
    # Count feedback per faculty
    faculty_counts = {}
    for f in feedbacks:
        fac_name = f.faculty.name
        faculty_counts[fac_name] = faculty_counts.get(fac_name, 0) + 1
    
    faculties = list(faculty_counts.keys())
    counts = list(faculty_counts.values())
    
    fig, ax = plt.subplots(figsize=(12, 6))
    colors = plt.cm.Blues(np.linspace(0.4, 0.8, len(faculties)))
    ax.barh(faculties, counts, color=colors, edgecolor='black')
    ax.set_xlabel('Number of Feedback Submissions', fontsize=12)
    ax.set_title('Feedback Count by Faculty', fontsize=14, fontweight='bold')
    ax.grid(axis='x', alpha=0.3)
    
    # Add value labels
    for i, (fac, count) in enumerate(zip(faculties, counts)):
        ax.text(count, i, f' {count}', va='center')
    
    img = BytesIO()
    fig.savefig(img, format='png', bbox_inches='tight')
    img.seek(0)
    plt.close(fig)
    
    return send_file(img, mimetype='image/png')


@app.route('/download/csv')
@login_required
def download_csv():
    if current_user.role != 'admin':
        return redirect(url_for('login'))
    feedbacks = Feedback.query.order_by(Feedback.created_at.desc()).all()
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['ID', 'Date', 'Faculty', 'Student Name', 'Teaching Quality', 'Course Content', 
                     'Communication', 'Feedback Quality', 'Subject Knowledge', 'Overall Rating', 'Comments'])
    for fb in feedbacks:
        writer.writerow([
            fb.id,
            fb.created_at.strftime('%Y-%m-%d %H:%M:%S') if fb.created_at else '',
            fb.faculty.name,
            fb.student_name or 'Anonymous',
            fb.q1_teaching_quality or '',
            fb.q2_course_content or '',
            fb.q3_communication or '',
            fb.q4_feedback_quality or '',
            fb.q5_subject_knowledge or '',
            fb.overall_rating or '',
            fb.comments or ''
        ])
    output.seek(0)
    return send_file(
        BytesIO(output.getvalue().encode('utf-8')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'feedback_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    )


@app.route('/download/excel')
@login_required
def download_excel():
    if current_user.role != 'admin':
        return redirect(url_for('login'))
    if not OPENPYXL_AVAILABLE:
        return "Error: openpyxl not installed. Run: py -m pip install openpyxl", 400
    
    feedbacks = Feedback.query.order_by(Feedback.created_at.desc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Feedback'
    
    # Header row
    headers = ['ID', 'Date', 'Faculty', 'Student Name', 'Teaching Quality', 'Course Content', 
               'Communication', 'Feedback Quality', 'Subject Knowledge', 'Overall Rating', 'Comments']
    ws.append(headers)
    
    # Format header
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for fb in feedbacks:
        ws.append([
            fb.id,
            fb.created_at.strftime('%Y-%m-%d %H:%M:%S') if fb.created_at else '',
            fb.faculty.name,
            fb.student_name or 'Anonymous',
            fb.q1_teaching_quality or '',
            fb.q2_course_content or '',
            fb.q3_communication or '',
            fb.q4_feedback_quality or '',
            fb.q5_subject_knowledge or '',
            fb.overall_rating or '',
            fb.comments or ''
        ])
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 40
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'feedback_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    # Use attachment_filename for compatibility with Flask versions
    try:
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except TypeError:
        # Fallback for older Flask versions
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            attachment_filename=filename
        )


if __name__ == '__main__':
    app.run(debug=True)
